import openpyxl
from datetime import datetime
from tabulate import tabulate
from openpyxl import load_workbook
from openpyxl.formula import Tokenizer

from rich import print
from rich import box
from rich.align import Align
from rich.panel import Panel

from rich.progress import track
from rich.progress import BarColumn, Progress, SpinnerColumn, TextColumn

from rich.traceback import install
install(show_locals=True)

# Load the Excel file
workbook = openpyxl.load_workbook('Examplery_data.xlsx')

# Select the active sheet
sheet = workbook.active

# Function to print the cell value of all months
def gross_profit_values():
    # Define the cell numbers for each month
    GROSS_PROFIT_month_cells = {
        'Jan': 'C6',
        'Feb': 'D6',
        'Mar': 'E6',
        'Apr': 'G6',
        'May': 'H6',
        'Jun': 'I6',
        'Jul': 'K6',
        'Aug': 'L6',
        'Sep': 'M6',
        'Oct': 'O6',
        'Nov': 'P6',
        'Dec': 'Q6'
    }

    panel_content = ""

    for month, cell_number in GROSS_PROFIT_month_cells.items():
        cell_value = sheet[cell_number].value
        panel_content += f"{month} = {cell_value}\n"

    panel_title = "Gross Profit Values"
    panel = Panel.fit(panel_content, title=panel_title, title_align="left", border_style="bold white", box=box.SQUARE)
    print(panel)

# Function to print the cell value of all months
def total_expenses_values():
    # Define the cell numbers for each month
    TOTAL_EXPENSES_month_cells = {
        'Jan': 'C7',
        'Feb': 'D7',
        'Mar': 'E7',
        'Apr': 'G7',
        'May': 'H7',
        'Jun': 'I7',
        'Jul': 'K7',
        'Aug': 'L7',
        'Sep': 'M7',
        'Oct': 'O7',
        'Nov': 'P7',
        'Dec': 'Q7'
    }

    panel_content = ""

    for month, cell_number in TOTAL_EXPENSES_month_cells.items():
        cell_value = sheet[cell_number].value
        panel_content += f"{month} = {cell_value}\n"

    panel_title = "Total Expenses"
    panel = Panel.fit(panel_content, title=panel_title, title_align="left", border_style="bold white", box=box.SQUARE)
    print(panel)

    
# Function to print the cell value of all months
def monthly_netprofit_values():
    # Define the cell numbers for each month
    MONTHY_NET_PROFIT_month_cells = {
        'Jan': 'C8',
        'Feb': 'D8',
        'Mar': 'E8',
        'Apr': 'G8',
        'May': 'H8',
        'Jun': 'I8',
        'Jul': 'K8',
        'Aug': 'L8',
        'Sep': 'M8',
        'Oct': 'O8',
        'Nov': 'P8',
        'Dec': 'Q8'
    }

    panel_content = ""

    # Load the workbook in read-only mode with data_only set to True to evaluate formulas
    wb = load_workbook('Examplery_data.xlsx', read_only=True, data_only=True)
    sheet = wb.active

    for month, cell_number in MONTHY_NET_PROFIT_month_cells.items():
        cell_value = sheet[cell_number].value

        panel_content += f"{month} = {cell_value}\n"

    panel_title = "Monthly Net Profit"
    panel = Panel.fit(panel_content, title=panel_title, title_align="left", border_style="bold white", box=box.SQUARE)
    print(panel)

# Function to print the cell value of all months
def year_to_date_netporfit():
    # Define the cell numbers for each month
    YEAR_TO_DATE_NETPROFIT_month_cells = {
        'Jan': 'C9',
        'Feb': 'D9',
        'Mar': 'E9',
        'Apr': 'G9',
        'May': 'H9',
        'Jun': 'I9',
        'Jul': 'K9',
        'Aug': 'L9',
        'Sep': 'M9',
        'Oct': 'O9',
        'Nov': 'P9',
        'Dec': 'Q9'
    }

    panel_content = ""

    # Load the workbook in read-only mode with data_only set to True to evaluate formulas
    wb = load_workbook('Examplery_data.xlsx', read_only=True, data_only=True)
    sheet = wb.active

    for month, cell_number in YEAR_TO_DATE_NETPROFIT_month_cells.items():
        cell_value = sheet[cell_number].value

        panel_content += f"{month} = {cell_value}\n"

    panel_title = "Year-to-date Net Profit"
    panel = Panel.fit(panel_content, title=panel_title, title_align="left", border_style="bold white", box=box.SQUARE)
    print(panel)

# Call the function to print the cell values of all months
# gross_profit_values()
# total_expenses_values()
# monthly_netprofit_values()
# year_to_date_netporfit()

# Functions for getting current and previous monthly values
def gross_profit_values_comparison(sheet):
    # Define the cell numbers for each month
    GROSS_PROFIT_month_cells = {
        'Jan': 'C6',
        'Feb': 'D6',
        'Mar': 'E6',
        'Apr': 'G6',
        'May': 'H6',
        'Jun': 'I6',
        'Jul': 'K6',
        'Aug': 'L6',
        'Sep': 'M6',
        'Oct': 'O6',
        'Nov': 'P6',
        'Dec': 'Q6'
    }

    def get_previous_month(current_month):
        months_list = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        current_index = months_list.index(current_month)
        previous_index = (current_index - 1) % len(months_list)
        return months_list[previous_index]

    current_month = datetime.now().strftime("%b")
    previous_month = get_previous_month(current_month)

    panel_content = ""

    current_value = sheet[GROSS_PROFIT_month_cells[current_month]].value
    previous_value = sheet[GROSS_PROFIT_month_cells[previous_month]].value

    panel_content += f"\nCurrent Month ({current_month}) Value = {current_value}\n"
    panel_content += f"Previous Month ({previous_month}) Value = {previous_value}\n"

    panel_title = "Gross Profit Values"
    panel = Panel.fit(panel_content, title=panel_title, title_align="left", border_style="bold white", box=box.SQUARE)
    print(panel)
    
# Functions for getting current and previous monthly values
def total_expenses_values_comparison(sheet):
    # Define the cell numbers for each month
    TOTAL_EXPENSES_month_cells = {
        'Jan': 'C7',
        'Feb': 'D7',
        'Mar': 'E7',
        'Apr': 'G7',
        'May': 'H7',
        'Jun': 'I7',
        'Jul': 'K7',
        'Aug': 'L7',
        'Sep': 'M7',
        'Oct': 'O7',
        'Nov': 'P7',
        'Dec': 'Q7'
    }

    def get_previous_month(current_month):
        months_list = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        current_index = months_list.index(current_month)
        previous_index = (current_index - 1) % len(months_list)
        return months_list[previous_index]

    current_month = datetime.now().strftime("%b")
    previous_month = get_previous_month(current_month)

    panel_content = ""

    current_value = sheet[TOTAL_EXPENSES_month_cells[current_month]].value
    previous_value = sheet[TOTAL_EXPENSES_month_cells[previous_month]].value

    panel_content += f"\nCurrent Month ({current_month}) Value = {current_value}\n"
    panel_content += f"Previous Month ({previous_month}) Value = {previous_value}\n"

    panel_title = "Total Expenses Values"
    panel = Panel.fit(panel_content, title=panel_title, title_align="left", border_style="bold white", box=box.SQUARE)
    print(panel)
    
# Functions for getting current and previous monthly values
def monthly_netprofit_values_comparison(sheet):
    # Define the cell numbers for each month
    MONTHY_NET_PROFIT_month_cells = {
        'Jan': 'C8',
        'Feb': 'D8',
        'Mar': 'E8',
        'Apr': 'G8',
        'May': 'H8',
        'Jun': 'I8',
        'Jul': 'K8',
        'Aug': 'L8',
        'Sep': 'M8',
        'Oct': 'O8',
        'Nov': 'P8',
        'Dec': 'Q8'
    }

    def get_previous_month(current_month):
        months_list = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        current_index = months_list.index(current_month)
        previous_index = (current_index - 1) % len(months_list)
        return months_list[previous_index]

    # Load the workbook in read-only mode with data_only set to True to evaluate formulas
    wb = load_workbook('Examplery_data.xlsx', read_only=True, data_only=True)
    sheet = wb.active

    current_month = datetime.now().strftime("%b")
    previous_month = get_previous_month(current_month)

    panel_content = ""

    current_value = sheet[MONTHY_NET_PROFIT_month_cells[current_month]].value
    previous_value = sheet[MONTHY_NET_PROFIT_month_cells[previous_month]].value

    panel_content += f"\nCurrent Month ({current_month}) Value = {current_value}\n"
    panel_content += f"Previous Month ({previous_month}) Value = {previous_value}\n"

    panel_title = "Monthly Net Profit"
    panel = Panel.fit(panel_content, title=panel_title, title_align="left", border_style="bold white", box=box.SQUARE)
    print(panel)

# Functions for getting current and previous monthly values
def ytd_netprofit_values_comparison(sheet):
    # Define the cell numbers for each month
    YEAR_TO_DATE_NET_PROFIT_month_cells = {
        'Jan': 'C9',
        'Feb': 'D9',
        'Mar': 'E9',
        'Apr': 'G9',
        'May': 'H9',
        'Jun': 'I9',
        'Jul': 'K9',
        'Aug': 'L9',
        'Sep': 'M9',
        'Oct': 'O9',
        'Nov': 'P9',
        'Dec': 'Q9'
    }

    def get_previous_month(current_month):
        months_list = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        current_index = months_list.index(current_month)
        previous_index = (current_index - 1) % len(months_list)
        return months_list[previous_index]

    # Load the workbook in read-only mode with data_only set to True to evaluate formulas
    wb = load_workbook('Examplery_data.xlsx', read_only=True, data_only=True)
    sheet = wb.active

    current_month = datetime.now().strftime("%b")
    previous_month = get_previous_month(current_month)

    panel_content = ""

    current_value = sheet[YEAR_TO_DATE_NET_PROFIT_month_cells[current_month]].value
    previous_value = sheet[YEAR_TO_DATE_NET_PROFIT_month_cells[previous_month]].value

    panel_content += f"\nCurrent Month ({current_month}) Value = {current_value}\n"
    panel_content += f"Previous Month ({previous_month}) Value = {previous_value}\n"

    panel_title = "Monthly Net Profit"
    panel = Panel.fit(panel_content, title=panel_title, title_align="left", border_style="bold white", box=box.SQUARE)
    print(panel)
    
gross_profit_values_comparison(sheet)
total_expenses_values_comparison(sheet)
monthly_netprofit_values_comparison(sheet)
ytd_netprofit_values_comparison(sheet)

def sales_values():
    # Define the cell numbers for each month
    SALES_month_cells = {
        'Jan': 'C19',
        'Feb': 'D19',
        'Mar': 'E19',
        'Apr': 'G19',
        'May': 'H19',
        'Jun': 'I19',
        'Jul': 'K19',
        'Aug': 'L19',
        'Sep': 'M19',
        'Oct': 'O19',
        'Nov': 'P19',
        'Dec': 'Q19'
    }

    panel_content = ""

    for month, cell_number in SALES_month_cells.items():
        cell_value = sheet[cell_number].value
        panel_content += f"{month} = {cell_value}\n"

    panel_title = "Sales of Products/Services Values"
    panel = Panel.fit(panel_content, title=panel_title, title_align="left", border_style="bold white", box=box.SQUARE)
    print(panel)
    
def commissions_values():
    # Define the cell numbers for each month
    COMMISSIONS_month_cells = {
        'Jan': 'C20',
        'Feb': 'D20',
        'Mar': 'E20',
        'Apr': 'G20',
        'May': 'H20',
        'Jun': 'I20',
        'Jul': 'K20',
        'Aug': 'L20',
        'Sep': 'M20',
        'Oct': 'O20',
        'Nov': 'P20',
        'Dec': 'Q20'
    }

    panel_content = ""

    for month, cell_number in COMMISSIONS_month_cells.items():
        cell_value = sheet[cell_number].value
        panel_content += f"{month} = {cell_value}\n"

    panel_title = "Commissions/Fees Values"
    panel = Panel.fit(panel_content, title=panel_title, title_align="left", border_style="bold white", box=box.SQUARE)
    print(panel)
    
def other_values():
    # Define the cell numbers for each month
    OTHER_month_cells = {
        'Jan': 'C21',
        'Feb': 'D21',
        'Mar': 'E21',
        'Apr': 'G21',
        'May': 'H21',
        'Jun': 'I21',
        'Jul': 'K21',
        'Aug': 'L21',
        'Sep': 'M21',
        'Oct': 'O21',
        'Nov': 'P21',
        'Dec': 'Q21'
    }

    panel_content = ""

    for month, cell_number in OTHER_month_cells.items():
        cell_value = sheet[cell_number].value
        panel_content += f"{month} = {cell_value}\n"

    panel_title = "Other Values"
    panel = Panel.fit(panel_content, title=panel_title, title_align="left", border_style="bold white", box=box.SQUARE)
    print(panel)
    
# Function to print the cell value of all months
def total_sales_values():

    # Define the cell numbers for each month
    TOTAL_SALES_month_cells = {
        'Jan': 'C22',
        'Feb': 'D22',
        'Mar': 'E22',
        'Apr': 'G22',
        'May': 'H22',
        'Jun': 'I22',
        'Jul': 'K22',
        'Aug': 'L22',
        'Sep': 'M22',
        'Oct': 'O22',
        'Nov': 'P22',
        'Dec': 'Q22'
    }

    panel_content = ""

    for month, cell_number in TOTAL_SALES_month_cells.items():
        cell = sheet[cell_number]

        # Check if the cell contains a formula
        if cell.data_type == "f":
            cell_value = cell.calculate()
        else:
            cell_value = cell.value

        # Convert the cell value to an integer if it's numeric, otherwise set it to 0
        if isinstance(cell_value, (int, float)):
            cell_value = int(cell_value)
        else:
            cell_value = 0

        panel_content += f"{month} = {cell_value}\n"

    panel_title = "Total Sales"
    panel = Panel.fit(panel_content, title=panel_title, title_align="left", border_style="bold white", box=box.SQUARE)
    print(panel)

print("-----------------------------------------------------------------------------")
sales_values()
commissions_values()
other_values()
total_sales_values()