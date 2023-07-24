from datetime import datetime
import openpyxl
from datetime import datetime
from tabulate import tabulate
from openpyxl import load_workbook
from openpyxl.formula import Tokenizer
import asciichartpy

from rich import print
from rich import box
from rich.tree import Tree
from rich.text import Text
from rich.align import Align
from rich.panel import Panel
from rich.layout import Layout 
from rich.table import Table

from rich.live import Live
from rich.prompt import Prompt
from rich.progress import track
from rich.progress import BarColumn, Progress, SpinnerColumn, TextColumn

from rich.traceback import install
install(show_locals=True)

layout = Layout()
# Load the Excel file
workbook = openpyxl.load_workbook('Examplery_data.xlsx')

# Select the active sheet
sheet = workbook.active



layout.split_column(
    Layout(name = "Header", size=3),
    Layout(name = "Body"),
    Layout(name = "Footer", size=3)
)

layout["Body"].split_column(
    Layout(name="Data", size=7),
    Layout(name="Charts")
)

layout["Data"].split_row(
    Layout(name="U1"),
    Layout(name="U2"),
    Layout(name="U3"),
    Layout(name="U4")
)


layout["Charts"].split_column(
    Layout(name="Upper_charts"),
    Layout(name="Lower_charts")
)

layout["Upper_charts"].split_row(
    Layout(name="Upper_chart1"),
    Layout(name="Upper_chart2")
)

layout["Lower_charts"].split_row(
    Layout(name="Lower_chart1"),
    Layout(name="Lower_chart2")
)

class Header:

    def __rich__(self) -> Panel:
        grid = Table.grid(expand=True)
        grid.add_column(justify="left")
        grid.add_column(justify="center", ratio=1)
        grid.add_column(justify="right")
        grid.add_row(
            "📝", "[b]Automate[/] - [i]Data Dashboard[/]", datetime.now().ctime().replace(":", "[blink]:[/]"),
        )
        return Panel(grid, style="bold white")
    
class Footer:

    def __rich__(self) -> Panel:
        grid = Table.grid(expand=True)
        grid.add_column(justify="center", ratio=1)
        grid.add_row("[i]Empowering Growth through Intelligent Automation[/]")
        return Panel(grid, style="white on black")
    
def gross_values(sheet):
    # Define the cell numbers for each month
    GROSS_PROFIT_month_cells = {
        'Jan': 'C6',
        'Feb': 'D6',
        'Mar': 'E6',
        'Apr': 'G6',
        'May': 'H6',
        'Jun': 'I6',
        'Jul': 'K6',
        'Aug': 'L6',
        'Sep': 'M6',
        'Oct': 'O6',
        'Nov': 'P6',
        'Dec': 'Q6'
    }

    def get_previous_month(current_month):
        months_list = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        current_index = months_list.index(current_month)
        previous_index = (current_index - 1) % len(months_list)
        return months_list[previous_index]

    current_month = datetime.now().strftime("%b")
    previous_month = get_previous_month(current_month)

    panel_content = ""

    current_value = sheet[GROSS_PROFIT_month_cells[current_month]].value
    previous_value = sheet[GROSS_PROFIT_month_cells[previous_month]].value

    panel_content += f"\nCurrent Month ({current_month}) Value = {current_value}\n"
    panel_content += f"Previous Month ({previous_month}) Value = {previous_value}\n"

    panel_title = "Gross Profit Values"
    panel = Panel.fit(panel_content, title=panel_title, title_align="left", border_style="bold white", box=box.SQUARE)
    return panel

def total_expenses(sheet):
    # Define the cell numbers for each month
    TOTAL_EXPENSES_month_cells = {
        'Jan': 'C7',
        'Feb': 'D7',
        'Mar': 'E7',
        'Apr': 'G7',
        'May': 'H7',
        'Jun': 'I7',
        'Jul': 'K7',
        'Aug': 'L7',
        'Sep': 'M7',
        'Oct': 'O7',
        'Nov': 'P7',
        'Dec': 'Q7'
    }

    def get_previous_month(current_month):
        months_list = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        current_index = months_list.index(current_month)
        previous_index = (current_index - 1) % len(months_list)
        return months_list[previous_index]

    current_month = datetime.now().strftime("%b")
    previous_month = get_previous_month(current_month)

    panel_content = ""

    current_value = sheet[TOTAL_EXPENSES_month_cells[current_month]].value
    previous_value = sheet[TOTAL_EXPENSES_month_cells[previous_month]].value

    panel_content += f"\nCurrent Month ({current_month}) Value = {current_value}\n"
    panel_content += f"Previous Month ({previous_month}) Value = {previous_value}\n"

    panel_title = "Total Expenses Values"
    panel = Panel.fit(panel_content, title=panel_title, title_align="left", border_style="bold white", box=box.SQUARE)
    return panel

def monthly_netprofits(sheet):
    # Define the cell numbers for each month
    MONTHY_NET_PROFIT_month_cells = {
        'Jan': 'C8',
        'Feb': 'D8',
        'Mar': 'E8',
        'Apr': 'G8',
        'May': 'H8',
        'Jun': 'I8',
        'Jul': 'K8',
        'Aug': 'L8',
        'Sep': 'M8',
        'Oct': 'O8',
        'Nov': 'P8',
        'Dec': 'Q8'
    }

    def get_previous_month(current_month):
        months_list = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        current_index = months_list.index(current_month)
        previous_index = (current_index - 1) % len(months_list)
        return months_list[previous_index]

    # Load the workbook in read-only mode with data_only set to True to evaluate formulas
    wb = load_workbook('Examplery_data.xlsx', read_only=True, data_only=True)
    sheet = wb.active

    current_month = datetime.now().strftime("%b")
    previous_month = get_previous_month(current_month)

    panel_content = ""

    current_value = sheet[MONTHY_NET_PROFIT_month_cells[current_month]].value
    previous_value = sheet[MONTHY_NET_PROFIT_month_cells[previous_month]].value

    panel_content += f"\nCurrent Month ({current_month}) Value = {current_value}\n"
    panel_content += f"Previous Month ({previous_month}) Value = {previous_value}\n"

    panel_title = "Monthly Net Profit"
    panel = Panel.fit(panel_content, title=panel_title, title_align="left", border_style="bold white", box=box.SQUARE)
    return panel

def ytd_netprofits(sheet):
    # Define the cell numbers for each month
    YEAR_TO_DATE_NET_PROFIT_month_cells = {
        'Jan': 'C9',
        'Feb': 'D9',
        'Mar': 'E9',
        'Apr': 'G9',
        'May': 'H9',
        'Jun': 'I9',
        'Jul': 'K9',
        'Aug': 'L9',
        'Sep': 'M9',
        'Oct': 'O9',
        'Nov': 'P9',
        'Dec': 'Q9'
    }

    def get_previous_month(current_month):
        months_list = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        current_index = months_list.index(current_month)
        previous_index = (current_index - 1) % len(months_list)
        return months_list[previous_index]

    # Load the workbook in read-only mode with data_only set to True to evaluate formulas
    wb = load_workbook('Examplery_data.xlsx', read_only=True, data_only=True)
    sheet = wb.active

    current_month = datetime.now().strftime("%b")
    previous_month = get_previous_month(current_month)

    panel_content = ""

    current_value = sheet[YEAR_TO_DATE_NET_PROFIT_month_cells[current_month]].value
    previous_value = sheet[YEAR_TO_DATE_NET_PROFIT_month_cells[previous_month]].value

    panel_content += f"\nCurrent Month ({current_month}) Value = {current_value}\n"
    panel_content += f"Previous Month ({previous_month}) Value = {previous_value}\n"

    panel_title = "Monthly Net Profit"
    panel = Panel.fit(panel_content, title=panel_title, title_align="left", border_style="bold white", box=box.SQUARE)
    return panel

def gross_profits_charts():
    # Define the cell numbers for each month
    GROSS_PROFIT_month_cells = {
        'Jan': 'C6',
        'Feb': 'D6',
        'Mar': 'E6',
        'Apr': 'G6',
        'May': 'H6',
        'Jun': 'I6',
        'Jul': 'K6',
        'Aug': 'L6',
        'Sep': 'M6',
        'Oct': 'O6',
        'Nov': 'P6',
        'Dec': 'Q6'
    }

    # Load the workbook in read-only mode with data_only set to True to evaluate formulas
    wb = load_workbook('Examplery_data.xlsx', read_only=True, data_only=True)
    sheet = wb.active

    # Initialize a list to store the net profit values
    gross_profit_values = []

    for month, cell_number in GROSS_PROFIT_month_cells.items():
        cell_value = sheet[cell_number].value
        gross_profit_values.append(cell_value)

    # Plot the graph using asciichartpy
    chart = asciichartpy.plot(gross_profit_values, {"width": 50, "height": 10, "format": "{:,.2f}"})
    return Panel(chart, title="Gross Profit values chart", box=box.SQUARE)

def monthly_netprofit_charts():
    # Define the cell numbers for each month
    MONTHLY_NET_PROFIT_month_cells = {
        'Jan': 'C8',
        'Feb': 'D8',
        'Mar': 'E8',
        'Apr': 'G8',
        'May': 'H8',
        'Jun': 'I8',
        'Jul': 'K8',
        'Aug': 'L8',
        'Sep': 'M8',
        'Oct': 'O8',
        'Nov': 'P8',
        'Dec': 'Q8'
    }

    # Load the workbook in read-only mode with data_only set to True to evaluate formulas
    wb = load_workbook('Examplery_data.xlsx', read_only=True, data_only=True)
    sheet = wb.active

    # Initialize a list to store the net profit values
    net_profit_values = []

    for month, cell_number in MONTHLY_NET_PROFIT_month_cells.items():
        cell_value = sheet[cell_number].value
        net_profit_values.append(cell_value)

    # Plot the graph using asciichartpy
    chart = asciichartpy.plot(net_profit_values, {"width": 50, "height": 10, "format": "{:,.2f}"})
    return Panel(chart, title="Monthly Net-Profit values chart", box=box.SQUARE)

layout["Header"].update(Header())
layout["Footer"].update(Footer())
layout["U1"].update(gross_values(sheet))
layout["U2"].update(total_expenses(sheet))
layout["U3"].update(monthly_netprofits(sheet))
layout["U4"].update(ytd_netprofits(sheet))
layout["Upper_chart1"].update(gross_profits_charts())
layout["Upper_chart2"].update(monthly_netprofit_charts())
print(layout)